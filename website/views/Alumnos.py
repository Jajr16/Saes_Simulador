import requests
from django.http import JsonResponse, HttpResponse
from django.views import View
from django.shortcuts import render, redirect
from .url import url
from ..forms import NAlumnoForm
from ..forms import NAlumnoVideoForm
from django.views.decorators.csrf import csrf_exempt
import os
import json

from openpyxl import load_workbook
import cloudinary
import cloudinary.uploader

from django.utils.decorators import method_decorator
from website.control_cargos import cargo_required

from django.http import FileResponse
import subprocess

def download_frame(request, boleta, frame_name):
    frame_path = f"/EntrenamientoIMG/{boleta}/{frame_name}"
    if os.path.exists(frame_path):
        return FileResponse(open(frame_path, 'rb'))
    else:
        return HttpResponse("Frame no encontrado", status=404)
    
def extract_frames(video_path, output_dir):
    try:
        os.makedirs(output_dir, exist_ok=True)
        ffmpeg_path = os.path.join(os.path.dirname(__file__), '..', 'ffmpeg', 'bin', 'ffmpeg')
        
        # Ejecutar FFmpeg (equivalente al ProcessBuilder de Java)
        subprocess.run([
            ffmpeg_path,
            "-i", video_path,
            "-vf", "fps=3",
            "-vsync", "vfr",
            f"{output_dir}/frame_%02d.png"
        ], check=True)
        
    except subprocess.CalledProcessError as e:
        raise Exception(f"Error al extraer frames: {e}")

@csrf_exempt
def obtener_imagen(request):
    if request.method == 'POST':
        try:
            print("Cuerpo de la solicitud:", request.body)
            # Obtener la ruta de la imagen desde la petición
            data = json.loads(request.body)
            ruta_imagen = data.get('ruta_imagen')

            if not ruta_imagen:
                return JsonResponse({'error': 'Ruta de imagen no proporcionada'}, status=400)

            #base_path = 'D:/Repositorio/Control-Acceso4/Control-Acceso/Programas/Saes_Sim/'
            #ruta_imagen_completa = os.path.join(base_path, ruta_imagen)

            print(f"Ruta de la imagen: {ruta_imagen}")

            if not os.path.exists(ruta_imagen):
                return JsonResponse({'error': 'Ruta de imagen no encontrada'}, status=404)

            # Leer la imagen y devolverla como respuesta
            with open(ruta_imagen, 'rb') as imagen_file:
                return HttpResponse(imagen_file.read(), content_type="image/jpeg")
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

def carreras(request):
    escuela = request.GET.get("escuela")
    
    api = "programasAcademicos"
    
    if not escuela:
        return JsonResponse({"error": "Falta el parámetro 'escuela'"}, status=400)

    try:
        escuela = int(escuela)
    except ValueError:
        return JsonResponse({"error": "El parámetro 'escuela' debe ser un número"}, status=400)

    headers = {"Content-Type": "application/json"}
    data = json.dumps({"escuela": escuela}) 

    try:
        response = requests.post(url+api, data=data, headers=headers)
        response.raise_for_status()
        response_data = response.json()
    except requests.exceptions.RequestException as e:
        return JsonResponse({"error": f"Error al conectar con la API: {str(e)}"}, status=500)

    return JsonResponse({"carreras": list(response_data)})

@method_decorator(cargo_required(allowed_roles=['Personal DAE']), name='dispatch') 
class AlumnoView(View):
    """
        Clase que define la vista del listado de los alumnos registrados
    """
    def get(self, request, *args, **kwargs):
        """
            Función get de la vista encargada de renderizar la página html
        """

        api = "alumnos"
        
        headers = {"Content-Type": "application/json"}
        
        response = requests.get(url+api, headers=headers)
        
        context = {
            "DetailAlumnos": response.json()
        }
        
        return render(request, 'Alumnos.html', context)
    
@method_decorator(cargo_required(allowed_roles=['Personal DAE']), name='dispatch') 
class NAlumnoView(View):
    """
        Clase que define la vista del formulario de alta de Alumnos
    """
    def get(self, request, *args, **kwargs):
        """
            Función para cargar la vista de la página html con el formulario
        """
        form = NAlumnoForm()
        return render(request, "New_Alumno.html", { 'form': form })
    
    def post(self, request, *args, **kwargs):
        """
            Función post para procesar los datos enviados del formulario
        """
        
        print(request)
        
        api = "nAlumno"
        form = NAlumnoForm(request.POST, request.FILES)
        
        if (form.is_valid()):
            curp = form.cleaned_data['curp']
            boleta = form.cleaned_data['boleta']
            nombre = form.cleaned_data['nombre']
            apellido_P = form.cleaned_data['apellido_P']
            apellido_M = form.cleaned_data['apellido_M']
            sexo = form.cleaned_data['sexo']
            correo = form.cleaned_data['correo']
            escuela = form.cleaned_data['escuela']
            carrera = form.cleaned_data['carrera']
            
            video = request.FILES.get("video-file")
            
            if not video:
                return render(request, 'New_Alumno.html', {'form': form, 'message': "El video es obligatorio", 'Error': True})
            
            # if credencial:
            #     with open(f"website/views/fotos/{boleta}.jpg", "wb") as f:
            #         for chunk in credencial.chunks():
            #             f.write(chunk)
                        
            # foto_path = f"website/views/fotos/{boleta}.jpg" if credencial else None
            
            files = {
                'video': request.FILES.get("video-file"),
                'credencial': request.FILES.get("foto-file"),
            }
            data = {
                "curp": curp,
                "boleta": boleta,
                "nombre": nombre,
                "apellido_p": apellido_P,
                "apellido_m": apellido_M,
                "sexo": sexo,
                "correo": correo,
                "escuela": int(escuela),
                "carrera": carrera,
            }
            
            response = requests.post(url+api, data=data, files=files)
            response_data = response.json()
            
            print(response_data)
            
            if response_data.get("Error"):
                return JsonResponse(response_data, status=400)
            else:
                data_red = {
                    "student_id": boleta,
                    "name": nombre,
                    "img_path": response_data.get("image_path")
                }
                response_red = requests.post("https://face-verification-app-bze0emevhsh5cvdz.mexicocentral-01.azurewebsites.net/api/register_student", data = data_red)
                response_red_data = response_red.json()
                print(f'La respuesta es {response_red_data}')
                
                if response_red_data.get("student_id") == boleta:
                    return JsonResponse(response_data, status=200)
                
                return JsonResponse({"Error": "Ocurrió un error al hacer el cálculo de la red neuronal."}, status=400)
        
        else:
            return JsonResponse({"message": form.errors, "Error": True}, status=400)

def asegurarse_de_crear_carpeta(carpeta):
    if not os.path.exists(carpeta):
        os.makedirs(carpeta)
        
@method_decorator(cargo_required(allowed_roles=['Personal DAE']), name='dispatch') 
class CargarAlumnoView(View):
    """
        Clase que define la vista del formulario del registro de fotos y registro del Alumno
    """
    def get(self, request, *args, **kwargs):
        """
            Función para cargar la vista con el formulario
        """
        form = NAlumnoVideoForm()
        return render(request, "New_Alumno_Video.html", { 'form': form })
        
        
    def post(self, request, *args, **kwargs):
        """
            Función post para procesar los datos enviados por el formulario
        """
        print(request)
        
        api = "nvAlumno"
        form = NAlumnoVideoForm(request.POST)
        
        if form.is_valid(): 
            boleta = form.cleaned_data['boleta']
            curp = form.cleaned_data['curp']
            
            # Define las rutas para las fotos y los videos
            foto_path = f"website/views/fotos/{boleta}.jpg"
            video_path = f"website/views/EntrenamientoIMG/{boleta}/{boleta}.mp4"
            frames_dir = f"website/views/EntrenamientoIMG/{boleta}"

            # Asegúrate de que las carpetas existen
            asegurarse_de_crear_carpeta(f"website/views/fotos")  # Carpeta para fotos
            asegurarse_de_crear_carpeta(frames_dir)  # Carpeta para frames

            # Procesar la foto
            credencial = request.FILES.get("foto-file")
            video = request.FILES.get("video-file")
            
            if not video:
                return render(request, 'New_Alumno_Video.html', {'form': form, 'message': "El video es obligatorio", 'Error': True})
            
            if credencial:
                with open(foto_path, "wb") as f:
                    for chunk in credencial.chunks():
                        f.write(chunk)
            
            # Guardar el video en la carpeta correspondiente
            with open(video_path, "wb") as f:
                for chunk in video.chunks():
                    f.write(chunk)

            # Extraer los frames del video
            extract_frames(video_path, frames_dir)

            # Lista de los nombres de los frames extraídos
            frames = [f"frame_{i:02d}.png" for i in range(1, len(os.listdir(frames_dir)) + 1)]

            # Realizar la llamada a la API
            data = {
                "boleta": boleta,
                "curp": curp,
                "credencial": foto_path
            }
            
            response = requests.post(url + api, data=data)
            response_data = response.json()
            
            print(response_data)

            if response_data.get("Error"):
                return JsonResponse(response_data, status=400)
            else:
                return JsonResponse(response_data, status=200)

        else:
            return JsonResponse({"message": "Error en el formulario", "Error": True}, status=400)
        
@method_decorator(cargo_required(allowed_roles=['Personal DAE']), name='dispatch') 
class CargarDatosView(View):
    """
        Clase que define la vista del formulario del registro de fotos y registro del Alumno
    """
    def get(self, request, *args, **kwargs):
        """
            Función para cargar la vista con el formulario
        """
        form = NAlumnoVideoForm()
        return render(request, "New_Alumno_Datos.html", { 'form': form })
        
        
    def post(self, request, *args, **kwargs):
        """
            Función post para procesar los datos enviados por el formulario
        """
        print(request)
        
        api = "nvAlumno"
        form = NAlumnoVideoForm(request.POST)
        
        if form.is_valid(): 
            boleta = form.cleaned_data['boleta']
            curp = form.cleaned_data['curp']
            
            # Define las rutas para las fotos y los videos
            foto_path = f"website/views/fotos/{boleta}.jpg"
            frames_dir = f"website/views/EntrenamientoIMG/{boleta}"
            excel_path = r"C:\\Users\\alfre\\OneDrive\\Documentos\\Semestres ESCOM\\Semestre 7 - ESCOM\\Control-Acceso\\Programas\\SpringBoot-Java\\src\\main\\java\\com\\example\\PruebaCRUD\\grupos_ESCOM.xlsx"

            boleta_en_excel = False

            wb = load_workbook(excel_path)
            try:
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(min_row=1, max_col=1):  # Solo primera columna
                        cell_value = str(row[0].value).strip() if row[0].value else ""
                        if cell_value == boleta:
                            boleta_en_excel = True
                            break
                    if boleta_en_excel:
                        break
            except Exception as e:
                return JsonResponse({"message": f"Error al leer el archivo Excel: {str(e)}", "Error": True}, status=500)
            finally:
                wb.close()

            if not os.path.exists(frames_dir) and not boleta_en_excel:
                return JsonResponse({"message": "La boleta no se encontró ni en las fotos ni en el Excel", "Error": True}, status=404)
            elif os.path.exists(frames_dir) and not boleta_en_excel:
                return JsonResponse({"message": "La boleta solo se encontró en las fotos", "Error": False}, status=200)
            elif not os.path.exists(frames_dir) and boleta_en_excel:
                return JsonResponse({"message": "La boleta solo se encontró en el Excel", "Error": False}, status=200)
            else: 
                
                credencial_url = None
                try:
                    if os.path.exists(foto_path):
                        upload_result = cloudinary.uploader.upload(foto_path, folder="fotosCredencial", public_id=boleta)
                        credencial_url = upload_result.get("secure_url")
                    else:
                        return JsonResponse({"message": "No se encontró la imagen de la credencial", "Error": True}, status=404)
                except Exception as e:
                    return JsonResponse({"message": f"Error al subir la credencial a Cloudinary: {str(e)}", "Error": True}, status=500)

                try:
                    if os.path.exists(frames_dir):
                        for filename in os.listdir(frames_dir):
                            frame_path = os.path.join(frames_dir, filename)
                            if os.path.isfile(frame_path) and filename.endswith(".png"):
                                cloudinary.uploader.upload(frame_path, folder=f"frames/{boleta}")
                    else:
                        return JsonResponse({"message": "No se encontró la carpeta de frames", "Error": True}, status=404)
                except Exception as e:
                    return JsonResponse({"message": f"Error al subir los frames a Cloudinary: {str(e)}", "Error": True}, status=500)

                # Realizar la llamada a la API
                data = {
                    "boleta": boleta,
                    "curp": curp,
                    "credencial": credencial_url
                }
                
                try:
                    response = requests.post(url + api, data=data)
                    response_data = response.json()
                except Exception as e:
                    return JsonResponse({"message": f"Error al conectar con la API externa: {str(e)}", "Error": True}, status=500)

                if response_data.get("Error"):
                    return JsonResponse(response_data, status=400)
                else:
                    return JsonResponse(response_data, status=200)
        else:
            return JsonResponse({"message": "Error en el formulario", "Error": True}, status=400)